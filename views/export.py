import csv
import io
import datetime as dt

from db.shifts import get_shifts
from utils import parse_period, shift_hours, lateness, hhmm_to_time, PERIOD_LABELS


def _dec(value: float) -> str:
    """Float → строка с запятой как разделителем (русский Excel)."""
    return f"{value:.2f}".replace(".", ",")


def render_csv(period: str, custom_from: str = "", custom_to: str = "",
               search: str = "") -> bytes:
    date_from, date_to = parse_period(period, custom_from, custom_to)
    shifts = get_shifts(date_from, date_to)
    if search:
        search_low = search.lower().strip()
        shifts = [s for s in shifts if search_low in s["worker_name"].lower()]

    # ── Сводка по работникам ─────────────────────────────────────────────────
    stats: dict[int, dict] = {}
    for s in shifts:
        wid = s["worker_id"]
        if wid not in stats:
            stats[wid] = {
                "name": s["worker_name"],
                "schedule": f"{s['default_start'] or '?'}–{s['default_end'] or '?'}",
                "shifts": 0,
                "hours": 0.0,
                "late": 0,
                "overtime": 0,
                "open": 0,
            }
        st = stats[wid]
        st["shifts"] += 1
        h = shift_hours(s)
        if h is not None:
            st["hours"] += h
        else:
            st["open"] += 1

        late_cls, _ = lateness(s)
        if late_cls in ("late", "very-late"):
            st["late"] += 1

        if s["left_at"] and s["default_end"]:
            d = dt.date.fromisoformat(s["date"])
            left = dt.datetime.fromisoformat(s["left_at"])
            sched_dt = dt.datetime.combine(d, hhmm_to_time(s["default_end"]))
            if (left - sched_dt).total_seconds() > 30 * 60:
                st["overtime"] += 1

    # ── Формируем CSV ────────────────────────────────────────────────────────
    buf = io.StringIO()
    wr = csv.writer(buf, delimiter=";")

    period_label = PERIOD_LABELS.get(period, period)
    df_str = date_from.strftime("%d.%m.%Y")
    dt_str = date_to.strftime("%d.%m.%Y")

    # Заголовок файла
    wr.writerow(["ЗАРЯД · Табель учёта рабочего времени"])
    wr.writerow(["Период:", f"{period_label}  {df_str} — {dt_str}"])
    if search:
        wr.writerow(["Фильтр:", search])
    wr.writerow([])

    # ── Раздел 1: сводка ─────────────────────────────────────────────────────
    wr.writerow(["СВОДКА ПО РАБОТНИКАМ"])
    wr.writerow([
        "ФИО", "График", "Смен", "Часов всего",
        "Среднее в день", "Опозданий", "Переработок", "Незакр. смен",
    ])

    total_shifts = 0
    total_hours = 0.0

    for st in sorted(stats.values(), key=lambda x: x["name"]):
        closed = st["shifts"] - st["open"]
        avg = round(st["hours"] / closed, 2) if closed > 0 else 0.0
        wr.writerow([
            st["name"],
            st["schedule"],
            st["shifts"],
            _dec(st["hours"]),
            _dec(avg),
            st["late"] or "",
            st["overtime"] or "",
            st["open"] or "",
        ])
        total_shifts += st["shifts"]
        total_hours += st["hours"]

    # Итоговая строка
    wr.writerow([
        "ИТОГО", "",
        total_shifts,
        _dec(total_hours),
        "", "", "", "",
    ])
    wr.writerow([])

    # ── Раздел 2: детализация по сменам ──────────────────────────────────────
    wr.writerow(["ДЕТАЛИЗАЦИЯ ПО СМЕНАМ"])
    wr.writerow([
        "Дата", "Работник", "График",
        "Приход", "Уход", "Часов",
        "Опоздание", "Статус",
    ])

    for s in sorted(shifts, key=lambda x: (x["worker_name"], x["date"])):
        arr = dt.datetime.fromisoformat(s["arrived_at"])
        left_str = ""
        if s["left_at"]:
            left_str = dt.datetime.fromisoformat(s["left_at"]).strftime("%H:%M")
        h = shift_hours(s)
        late_cls, late_lbl = lateness(s)
        if not s["left_at"]:
            status = "открыта"
        elif s["auto_closed"]:
            status = "авто"
        else:
            status = ""
        wr.writerow([
            arr.strftime("%d.%m.%Y"),
            s["worker_name"],
            f"{s['default_start'] or '?'}–{s['default_end'] or '?'}",
            arr.strftime("%H:%M"),
            left_str,
            _dec(h) if h is not None else "",
            late_lbl if late_cls else "",
            status,
        ])

    return ("﻿" + buf.getvalue()).encode("utf-8")


def render_xlsx(period: str, custom_from: str = "", custom_to: str = "",
                search: str = "") -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl не установлен. Запусти: pip install openpyxl")

    import io

    date_from, date_to = parse_period(period, custom_from, custom_to)
    shifts = get_shifts(date_from, date_to)
    if search:
        search_low = search.lower().strip()
        shifts = [s for s in shifts if search_low in s["worker_name"].lower()]

    # Build stats (same logic as render_csv)
    stats: dict[int, dict] = {}
    for s in shifts:
        wid = s["worker_id"]
        if wid not in stats:
            stats[wid] = {
                "name": s["worker_name"],
                "schedule": f"{s['default_start'] or '?'}–{s['default_end'] or '?'}",
                "shifts": 0, "hours": 0.0, "late": 0, "overtime": 0, "open": 0,
            }
        st = stats[wid]
        st["shifts"] += 1
        h = shift_hours(s)
        if h is not None:
            st["hours"] += h
        else:
            st["open"] += 1
        late_cls, _ = lateness(s)
        if late_cls in ("late", "very-late"):
            st["late"] += 1
        if s["left_at"] and s["default_end"]:
            d = dt.date.fromisoformat(s["date"])
            left = dt.datetime.fromisoformat(s["left_at"])
            sched_dt = dt.datetime.combine(d, hhmm_to_time(s["default_end"]))
            if (left - sched_dt).total_seconds() > 30 * 60:
                st["overtime"] += 1

    wb = Workbook()
    ws = wb.active
    ws.title = "Табель"

    fill_header = PatternFill(fill_type="solid", fgColor="2D333B")
    fill_total = PatternFill(fill_type="solid", fgColor="FFD60A")
    bold = Font(bold=True)
    white_bold = Font(bold=True, color="FFFFFF")
    dark_bold = Font(bold=True, color="0D1117")

    period_label = PERIOD_LABELS.get(period, period)
    df_str = date_from.strftime("%d.%m.%Y")
    dt_str = date_to.strftime("%d.%m.%Y")

    r = 1
    ws.cell(r, 1, "ЗАРЯД · Табель учёта рабочего времени").font = Font(bold=True, size=13)
    ws.merge_cells(f"A{r}:H{r}")
    r += 1
    ws.cell(r, 1, "Период:").font = bold
    ws.cell(r, 2, f"{period_label}  {df_str} — {dt_str}")
    r += 1
    if search:
        ws.cell(r, 1, "Фильтр:").font = bold
        ws.cell(r, 2, search)
        r += 1
    r += 1

    # --- Summary ---
    ws.cell(r, 1, "СВОДКА ПО РАБОТНИКАМ").font = Font(bold=True, size=11)
    ws.merge_cells(f"A{r}:H{r}")
    r += 1
    for col, h in enumerate(
        ["ФИО", "График", "Смен", "Часов всего", "Среднее/день", "Опозданий", "Переработок", "Незакр."], 1
    ):
        cell = ws.cell(r, col, h)
        cell.font = white_bold
        cell.fill = fill_header
    r += 1

    total_shifts = 0
    total_hours = 0.0
    for st in sorted(stats.values(), key=lambda x: x["name"]):
        closed = st["shifts"] - st["open"]
        avg = round(st["hours"] / closed, 2) if closed > 0 else 0.0
        for col, val in enumerate([
            st["name"], st["schedule"], st["shifts"],
            round(st["hours"], 2), avg,
            st["late"] or "", st["overtime"] or "", st["open"] or "",
        ], 1):
            ws.cell(r, col, val)
        total_shifts += st["shifts"]
        total_hours += st["hours"]
        r += 1

    for col, val in enumerate(["ИТОГО", "", total_shifts, round(total_hours, 2), "", "", "", ""], 1):
        cell = ws.cell(r, col, val)
        cell.font = dark_bold
        cell.fill = fill_total
    r += 2

    # --- Detail ---
    ws.cell(r, 1, "ДЕТАЛИЗАЦИЯ ПО СМЕНАМ").font = Font(bold=True, size=11)
    ws.merge_cells(f"A{r}:H{r}")
    r += 1
    for col, h in enumerate(
        ["Дата", "Работник", "График", "Приход", "Уход", "Часов", "Опоздание", "Статус"], 1
    ):
        cell = ws.cell(r, col, h)
        cell.font = white_bold
        cell.fill = fill_header
    r += 1

    for s in sorted(shifts, key=lambda x: (x["worker_name"], x["date"])):
        arr = dt.datetime.fromisoformat(s["arrived_at"])
        left_str = dt.datetime.fromisoformat(s["left_at"]).strftime("%H:%M") if s["left_at"] else ""
        h = shift_hours(s)
        late_cls, late_lbl = lateness(s)
        status = "открыта" if not s["left_at"] else ("авто" if s["auto_closed"] else "")
        for col, val in enumerate([
            arr.strftime("%d.%m.%Y"), s["worker_name"],
            f"{s['default_start'] or '?'}–{s['default_end'] or '?'}",
            arr.strftime("%H:%M"), left_str,
            round(h, 2) if h is not None else "",
            late_lbl if late_cls else "", status,
        ], 1):
            ws.cell(r, col, val)
        r += 1

    for i, width in enumerate([18, 24, 14, 10, 10, 10, 14, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
